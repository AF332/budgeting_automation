from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


BUDGET_FILE = Path(__file__).parent.parent / "Simple personal budget.xlsx"

MONTH_TO_COL = {
    1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8,
    7: 9, 8: 10, 9: 11, 10: 12, 11: 13, 12: 14,
}

INCOME_ROWS = {
    "wages":         6,
    "reimbursement": 7,
    "miscellaneous": 8,
}

EXPENSE_ROWS = {
    "rent":                  13,
    "home_insurance":        14,
    "home_bits":             15,
    "services":              16,
    "utilities":             17,
    "groceries":             21,
    "clothes":               22,
    "dry_cleaning":          23,
    "dining_out":            24,
    "housecleaning":         25,
    "dog_walker":            26,
    "flight":                30,
    "transport_insurance":   31,
    "repairs":               32,
    "car_wash":              33,
    "uber":                  34,
    "public_transport":      35,
    "phone":                 39,
    "gaming":                40,
    "movies":                41,
    "concerts":              42,
    "gym":                   65,
    "sports_equipment":      66,
    "supplements":           67,
    "disney":                72,
    "chatgpt":               77,
    "charity":               78,
    "steam":                 82,
    "gifts":                 83,
    "barber":                84,
    "books":                 85,
    "savings":               90,
    "investment":            91,
    "credit_card":           92,
    "vape":                  98,
    "amazon":               101,
    "birthdays":            102,
}

UNCATEGORISED_ROW = 94

DESCRIPTION_RULES: list[tuple[str, str]] = [
    ("tesco|asda|lidl|aldi|sainsbury|morrisons|waitrose|farmfoods|home bargains|b&m|poundland|food|grocery", "groceries"),
    ("mcdonald|kfc|subway|pizza|kebab|nando|wagamama|burger|restaurant|cafe|coffee|starbucks|costa|greggs|dining|eating|takeaway|deliveroo|just eat|uber eats|marmaris", "dining_out"),
    ("rent|propsync|landlord|letting", "rent"),
    ("electric|gas|water|british gas|octopus|edf|eon|bulb|utility|utilities", "utilities"),
    ("broadband|virgin media|bt|sky|talktalk|internet", "services"),
    ("h3g|three|vodafone|ee|o2|giffgaff|phone bill|mobile", "phone"),
    ("pure gym|gym|fitness|leisure centre", "gym"),
    ("trainline|train|avanti|lner|gwr|tfl|bus|coach|national rail|stagecoach|arriva|megabus|flixbus|public transport", "public_transport"),
    ("uber|bolt|lyft|taxi|cab", "uber"),
    ("ryanair|easyjet|jet2|british airways|flight|airport", "flight"),
    ("disney|netflix|spotify|apple music|youtube premium|twitch|prime video|now tv|streaming", "disney"),
    ("chatgpt|openai", "chatgpt"),
    ("charity|donation|oxfam|red cross|islamic relief|comic relief", "charity"),
    ("steam|epic games|playstation|xbox|nintendo|gaming|valorant|game", "gaming"),
    ("amazon|amzn", "amazon"),
    ("barber|hairdress|hair cut|salon|handr", "barber"),
    ("savings|saving|pot_|pot |piggybank|investment|trading 212|freetrade", "savings"),
    ("birthday|present|gift|card factory", "birthdays"),
    ("salary|wages|payroll|digital futures|employer|income", "wages"),
    ("reimbursement|reimburse|refund", "reimbursement"),
    ("vape|vaping|lost mary|elf bar", "vape"),
    ("book|kindle|waterstones", "books"),
    ("supplement|protein|myprotein|bulk|vitamin", "supplements"),
]

CATEGORY_API_MAP: dict[str, str] = {
    "groceries":     "groceries",
    "eating_out":    "dining_out",
    "transport":     "public_transport",
    "entertainment": "movies",
    "income":        "wages",
    "savings":       "savings",
    "bills":         "utilities",
    "general":       "",
    "transfer":      "",
}


def categorise_transaction(description: str, api_category: str) -> str:
    """
    Determines the budget row key for a transaction using its description
    and API-assigned category.

    First attempts to match the description against a prioritised list of
    keyword rules. If no match is found, falls back to the API category
    mapping. Returns an empty string if the transaction cannot be categorised,
    which causes it to be written to the uncategorised row.

    Args:
        description (str): The transaction description or merchant name.
        api_category (str): The category label assigned by the bank API
            e.g. 'groceries', 'eating_out', 'transport'.

    Returns:
        str: A budget row key matching a key in INCOME_ROWS or EXPENSE_ROWS,
        or an empty string if uncategorised.
    """
    desc_lower = description.lower()
    for pattern, budget_key in DESCRIPTION_RULES:
        if re.search(pattern, desc_lower):
            return budget_key
    return CATEGORY_API_MAP.get(api_category.lower(), "")


def get_sheet_name(year: int) -> str:
    """
    Returns the Excel sheet name for the given year.

    Args:
        year (int): The year being processed e.g. 2026.

    Returns:
        str: The sheet name e.g. 'PERSONAL BUDGET 2026'.
    """
    return f"PERSONAL BUDGET {year}"


def write_transactions_to_budget(
    transactions: list[dict],
    year: int,
    month: int,
) -> Path:
    """
    Writes a list of normalised transactions directly into the master budget
    Excel file in the correct month column for the given year.

    Rather than creating a copy, this function loads the master budget file,
    writes the monthly totals into the appropriate cells, and saves it back
    in place. Each month's data goes into its own column so running the
    script for January fills column C, February fills column D, and so on.
    Existing data in other month columns is never touched.

    Groups all transactions by budget category, sums the amounts for each
    category, and writes the totals in pounds as plain numeric values. The
    existing SUBTOTAL and SUM formulas in the sheet recalculate automatically
    when the file is opened in Excel. Uncategorised transactions are summed
    and written to the 'Other obligations' row.

    Also creates or replaces a per-month audit sheet listing every individual
    transaction with date, description, amount, category, and source bank.
    Uncategorised transactions are highlighted in amber.

    Args:
        transactions (list[dict]): A list of normalised transaction
            dictionaries each containing date, description, amount (pence),
            category, and source fields.
        year (int): The calendar year being processed e.g. 2026.
        month (int): The calendar month being processed as an integer from
            1 to 12.

    Returns:
        Path: The path to the master budget file that was updated.

    Raises:
        FileNotFoundError: If the master budget file does not exist at the
            expected path.
        ValueError: If the sheet for the given year does not exist in the
            workbook.
    """
    if not BUDGET_FILE.exists():
        raise FileNotFoundError(
            f"Budget file not found at {BUDGET_FILE}. "
            "Make sure Simple_personal_budget.xlsx is in your project folder."
        )

    wb = load_workbook(BUDGET_FILE)
    sheet_name = get_sheet_name(year)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    col = MONTH_TO_COL[month]
    month_name = datetime(year, month, 1).strftime("%B")

    income_totals: dict[str, float] = {k: 0.0 for k in INCOME_ROWS}
    expense_totals: dict[str, float] = {k: 0.0 for k in EXPENSE_ROWS}
    uncategorised_total: float = 0.0
    uncategorised_items: list[dict] = []

    for t in transactions:
        pounds = abs(t["amount"]) / 100
        is_income = t["amount"] > 0
        budget_key = categorise_transaction(t["description"], t["category"])

        if is_income:
            if budget_key in income_totals:
                income_totals[budget_key] += pounds
            else:
                income_totals["miscellaneous"] += pounds
        else:
            if budget_key in expense_totals:
                expense_totals[budget_key] += pounds
            else:
                uncategorised_total += pounds
                uncategorised_items.append(t)

    for key, row in INCOME_ROWS.items():
        total = income_totals[key]
        if total > 0:
            ws.cell(row=row, column=col).value = round(total, 2)
            ws.cell(row=row, column=col).font = Font(color="000000", size=11)

    for key, row in EXPENSE_ROWS.items():
        total = expense_totals[key]
        if total > 0:
            ws.cell(row=row, column=col).value = round(total, 2)
            ws.cell(row=row, column=col).font = Font(color="000000", size=11)

    if uncategorised_total > 0:
        current = ws.cell(row=UNCATEGORISED_ROW, column=col).value
        try:
            existing = float(current) if current is not None else 0.0
        except (ValueError, TypeError):
            existing = 0.0
        ws.cell(row=UNCATEGORISED_ROW, column=col).value = round(
            existing + uncategorised_total, 2
        )
        ws.cell(row=UNCATEGORISED_ROW, column=col).font = Font(color="000000", size=11)

    _write_audit_sheet(wb, transactions, uncategorised_items, year, month)

    wb.save(BUDGET_FILE)

    print(f"\n  {month_name} {year} written to {BUDGET_FILE.name}")
    print(f"  Income rows updated:  {sum(1 for v in income_totals.values() if v > 0)}")
    print(f"  Expense rows updated: {sum(1 for v in expense_totals.values() if v > 0)}")
    if uncategorised_items:
        print(f"  Uncategorised:        {len(uncategorised_items)} transactions (see audit sheet)")

    return BUDGET_FILE


def _write_audit_sheet(
    wb,
    transactions: list[dict],
    uncategorised: list[dict],
    year: int,
    month: int,
) -> None:
    """
    Creates or replaces the per-month transactions audit sheet in the
    workbook with a full log of every transaction processed.

    The sheet is named 'Transactions YYYY-MM' so each month gets its own
    tab and previous months are never overwritten. Contains one row per
    transaction with columns for date, description, amount, budget category,
    source bank, and status. Income is formatted in green, spending in red,
    and uncategorised transactions are highlighted in amber.

    Args:
        wb: An openpyxl Workbook object.
        transactions (list[dict]): All normalised transactions for the month.
        uncategorised (list[dict]): Subset of transactions that could not
            be categorised automatically.
        year (int): The calendar year being processed.
        month (int): The calendar month being processed.

    Returns:
        None
    """
    sheet_name = f"Transactions {year}-{month:02d}"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    headers = ["Date", "Description", "Amount (£)", "Category", "Source", "Status"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4F4F")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 16

    uncategorised_descs = {
        (t["description"], t["amount"]) for t in uncategorised
    }

    green_font = Font(color="006400")
    red_font = Font(color="8B0000")
    amber_fill = PatternFill("solid", fgColor="FFF3CD")
    thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))

    for row_idx, t in enumerate(transactions, 2):
        pounds = t["amount"] / 100
        budget_key = categorise_transaction(t["description"], t["category"])
        is_uncategorised = (
            (t["description"], t["amount"]) in uncategorised_descs
            and t["amount"] < 0
        )
        status = "Uncategorised" if is_uncategorised else "OK"

        values = [
            t["date"],
            t["description"],
            round(pounds, 2),
            budget_key or t["category"],
            t["source"],
            status,
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if col_idx == 3:
                cell.number_format = '£#,##0.00;[Red](£#,##0.00)'
                cell.font = green_font if pounds > 0 else red_font

            if is_uncategorised:
                cell.fill = amber_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(transactions) + 1}"

    if uncategorised:
        print(f"\n  WARNING: {len(uncategorised)} uncategorised transactions (highlighted in amber in '{sheet_name}' tab):")
        for t in uncategorised[:10]:
            print(f"    {t['date']} | £{abs(t['amount'])/100:.2f} | {t['description']}")
        if len(uncategorised) > 10:
            print(f"    ... and {len(uncategorised) - 10} more.")